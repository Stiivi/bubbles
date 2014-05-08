from itertools import dropwhile

from bubbles.errors import ArgumentError, BubblesError
from bubbles.metadata import FieldList, Field
from bubbles.objects import DataObject
from bubbles.stores import DataStore

from openpyxl import load_workbook, Workbook
from openpyxl.cell import Cell


CELL_TYPES = {Cell.TYPE_STRING: 'string',
              Cell.TYPE_FORMULA: 'unknown',
              Cell.TYPE_NUMERIC: 'float',
              Cell.TYPE_BOOL: 'boolean',
              Cell.TYPE_NULL: 'unknown',
              Cell.TYPE_INLINE: 'string',
              Cell.TYPE_ERROR: 'unknown',
              Cell.TYPE_FORMULA_CACHE_STRING: 'string'}


def _load_workbook(resource):
    return load_workbook(resource, use_iterators=True,
                         keep_vba=False, data_only=True)


class XLSXStore(DataStore):
    def __init__(self, resource, encoding=None):
        super(XLSXStore, self).__init__()

        self.resource = resource
        self.encoding = encoding
        self._book = None

    @property
    def book(self):
        if self._book is None:
            self._book = _load_workbook(self.resource)
        return self._book

    def get_object(self, name, skip_rows=0, has_header=True):
        return XLSXObject(self.book, sheet=name, encoding=self.encoding,
                          skip_rows=skip_rows, has_header=has_header)

    def object_names(self):
        return self.book.get_sheet_names()

    def create(self, name):
        raise BubblesError("XLSX store is read-only")


class XLSXObject(DataObject):
    __identifier__ = "xlsx"

    def __init__(self, resource=None, fields=None, sheet=0,
                 encoding=None, skip_rows=0, has_header=True):
        """Creates a XLSX spreadsheet data source stream.

        Attributes:

        * resource: file name, URL or file-like object
        * sheet: sheet index number (as int) or sheet name
        * has_header: flag determining whether first line contains header or
          not. ``True`` by default.
        """
        if isinstance(resource, Workbook):
            self.workbook = resource
        else:
            self.workbook = _load_workbook(resource)

        if isinstance(sheet, int):
            self.sheet = self.workbook.worksheets[sheet]
        elif isinstance(sheet, str):
            self.sheet = self.workbook[sheet]
        else:
            raise ArgumentError('sheet has to be a string or an integer')

        if has_header:
            self.first_row = skip_rows + 1
        else:
            self.first_row = skip_rows

        if fields:
            self.fields = fields
        else:
            rows = enumerate(self.sheet.rows)
            first_row = next(dropwhile(lambda x: x[0] < self.first_row,
                                       rows))[1]
            if has_header:
                header_row = next(self.sheet.rows)
                names = [str(c.value) for c in header_row]
            else:
                names = ['col%d' % i for i in range(len(first_row))]

            self.fields = FieldList()
            for name, cell in zip(names, first_row):
                if cell.is_date:
                    storage_type = 'date'
                else:
                    storage_type = CELL_TYPES.get(cell.data_type, 'unknown')
                field = Field(name, storage_type=storage_type)
                self.fields.append(field)

    def _nrows(self):
        return sum([1 for _ in self.sheet.rows])

    def representations(self):
        return ["rows", "records"]

    def __len__(self):
        return self._nrows() - self.first_row

    def rows(self):
        if not self.fields:
            raise RuntimeError("Fields are not initialized")

        field_count = len(self.fields)
        rows = enumerate(self.sheet.rows)
        rows = dropwhile(lambda x: x[0] < self.first_row, rows)
        for _, row in rows:
            yield tuple([c.value for c in row[:field_count]])
